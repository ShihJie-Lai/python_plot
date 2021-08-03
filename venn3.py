import sys
from tkinter import *
from tkinter.filedialog import askopenfilename
import cv2
import re
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from matplotlib_venn import venn3, venn3_circles
from matplotlib import pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors, Font, Fill, NamedStyle,Color
from upsetplot import from_memberships
from upsetplot import plot



def venn_d(xlsx1="", xlsx2="", xlsx3="",os_s="",n_1="",n_2="",n_3=""):
	
	print(xlsx1)
	print(xlsx2)
	print(xlsx3)
	######
	ft1=Font(name='Arial',color=colors.RED, size=12,b=True)
	ali = Alignment(horizontal='center', vertical='center')
	######

	#print(sheet['A1'].value)
	wo = Workbook()
	ws1 = wo.create_sheet('ALL',index=0)
	ws2 = wo.create_sheet('intersection',index=1)
	ws3 = wo.create_sheet('Group1_only',index=2)
	ws4 = wo.create_sheet('Group2_only',index=3)
	ws5 = wo.create_sheet('Group3_only',index=4)
	ws6 = wo.create_sheet('Group1_Group2_only',index=5)
	ws7 = wo.create_sheet('Group1_Group3_only',index=6)
	ws8 = wo.create_sheet('Group2_Group3_only',index=7)
	#########excel_1#########
	wb = load_workbook(xlsx1)
	sheets = wb.sheetnames
	sheet1 = wb.get_sheet_by_name(sheets[0])
	sheet1_1 = wb.get_sheet_by_name(sheets[1])
	a=[]
	for i in range(0,sheet1_1.max_column):
		a.append([])
	ii=0
	for col in range(1,sheet1_1.max_column+1):
		for row in range(2,sheet1_1.max_row+1):
			a1=sheet1_1.cell(row =row,column = col).value
			#ws1.cell(row =row+1,column = col).value=a1
			#ws1.cell(row =3,column = col).font=ft1
			if(row==2):
				ws2.cell(row=3,column=col).value=a1
				ws2.cell(row=3,column=col).font=ft1
				ws3.cell(row=2,column=col).value=a1
				ws3.cell(row=2,column=col).font=ft1
				ws6.cell(row=3,column=col).value=a1
				ws6.cell(row=3,column=col).font=ft1
				ws7.cell(row=3,column=col).value=a1
				ws7.cell(row=3,column=col).font=ft1
			a[ii].append(a1)
		ii=ii+1
	a_1=[]
	for row in range(3,sheet1_1.max_row+1):
		a1=sheet1_1.cell(row =row,column = 1).value
		a_1.append(a1)
	###########excel_2#########
	wb1 = load_workbook(xlsx2)
	sheets = wb1.sheetnames
	sheet = wb1.get_sheet_by_name(sheets[0])
	sheet_1 = wb1.get_sheet_by_name(sheets[1])
	b=[]
	for i in range(0,sheet_1.max_column):
		b.append([])
	ii=0
	for col in range(1,sheet_1.max_column+1):
		for row in range(2,sheet_1.max_row+1):
			a1=sheet_1.cell(row =row,column = col).value
			b[ii].append(a1)
			if(row==2):
				ws4.cell(row=2,column=col).value=a1
				ws4.cell(row=2,column=col).font=ft1
				ws8.cell(row=3,column=col).value=a1
				ws8.cell(row=3,column=col).font=ft1
			if(col>7):
				# ws1.cell(row =row+1,column = sheet1.max_column+col-7).value=a1
				# ws1.cell(row =3,column = sheet1.max_column+col-7).font=ft1
				if(row==2):
					ws2.cell(row=3,column=sheet1.max_column+col-7).value=a1
					ws2.cell(row=3,column=sheet1.max_column+col-7).font=ft1
					ws6.cell(row=3,column=sheet1.max_column+col-7).value=a1
					ws6.cell(row=3,column=sheet1.max_column+col-7).font=ft1
		ii=ii+1
	b_1=[]
	for row in range(3,sheet_1.max_row+1):
		a1=sheet_1.cell(row =row,column = 1).value
		b_1.append(a1)#list 
		
	# #########excel_3#########
	wb2 = load_workbook(xlsx3)
	sheets = wb2.sheetnames
	sheet2 = wb2.get_sheet_by_name(sheets[0])
	sheet2_1 = wb2.get_sheet_by_name(sheets[1])
	c=[]
	for i in range(0,sheet2_1.max_column):
		c.append([])
	ii=0
	for col in range(1,sheet2_1.max_column+1):
		for row in range(2,sheet2_1.max_row+1):
			a1=sheet2_1.cell(row =row,column = col).value
			c[ii].append(a1)
			if(row==2):
				ws5.cell(row=2,column=col).value=a1
				ws5.cell(row=2,column=col).font=ft1
			if(col>7):
				# ws1.cell(row =row+1,column = sheet1.max_column+col-7).value=a1
				# ws1.cell(row =3,column = sheet1.max_column+col-7).font=ft1
				if(row==2):
					ws2.cell(row=3,column=sheet1.max_column+sheet.max_column+col-7-7).value=a1
					ws2.cell(row=3,column=sheet1.max_column+sheet.max_column+col-7-7).font=ft1
					ws7.cell(row=3,column=sheet1.max_column+col-7).value=a1
					ws7.cell(row=3,column=sheet1.max_column+col-7).font=ft1
					ws8.cell(row=3,column=sheet1.max_column+col-7).value=a1
					ws8.cell(row=3,column=sheet1.max_column+col-7).font=ft1
		ii=ii+1
	c_1=[]
	for row in range(3,sheet2_1.max_row+1):
		a1=sheet2_1.cell(row =row,column = 1).value
		c_1.append(a1)#list 		
		
		
	##########intersection##############
	inte = list(set(a_1) & (set(b_1)) & (set(c_1)) )
	le1=len(inte)
	for row11 in inte:
		ss_a=[]
		ss_b=[]
		ss_c=[]
		d=a[0].index(row11)
		e=b[0].index(row11)
		f=c[0].index(row11)
		for cell in range(0,(sheet.max_column)):

			results =a[cell][d]
			ss_a.append(results)
			if(cell>6):
				results =b[cell][e]
				ss_b.append(results)
				results =c[cell][f]
				ss_c.append(results)
		ss_a.extend(ss_b)
		ss_a.extend(ss_c)
		ws2.append(ss_a)
	##########Group1##############	
	diff1 = list(set(a_1)-(set(b_1))-(set(c_1)))
	le2=len(diff1)
	for row11 in diff1:
		ss_a=[]
		d=a[0].index(row11)
		for cell in range(0,(sheet.max_column)):
			results =a[cell][d]
			ss_a.append(results)
		ws3.append(ss_a)	
	##########Group2##############	
	diff1 = list(set(b_1)-(set(a_1))-(set(c_1)))
	le3=len(diff1)
	for row11 in diff1:
		ss_b=[]
		d=b[0].index(row11)
		for cell in range(0,(sheet.max_column)):
			results =b[cell][d]
			ss_b.append(results)
		ws4.append(ss_b)
		
	##########Group3##############	
	diff1 = list(set(c_1)-(set(a_1))-(set(b_1)))
	le4=len(diff1)
	for row11 in diff1:
		ss_c=[]
		d=c[0].index(row11)
		for cell in range(0,(sheet.max_column)):
			results =c[cell][d]
			ss_c.append(results)
		ws5.append(ss_c)	
	##########Group1-Group2##############	
	diff1 = list((set(a_1)) & (set(b_1)) - (set(c_1)))
	le5=len(diff1)
	for row11 in diff1:
		ss_a=[]
		ss_b=[]
		d=a[0].index(row11)
		e=b[0].index(row11)
		for cell in range(0,(sheet.max_column)):
			results =a[cell][d]
			ss_a.append(results)
			if(cell>6):
				results =b[cell][e]
				ss_b.append(results)
		ss_a.extend(ss_b)
		ws6.append(ss_a)
	##########Group1-Group3##############	
	diff1 = list((set(a_1)) & (set(c_1)) - (set(b_1)))
	le6=len(diff1)
	for row11 in diff1:
		ss_a=[]
		ss_c=[]
		d=a[0].index(row11)
		e=c[0].index(row11)
		for cell in range(0,(sheet.max_column)):
			results =a[cell][d]
			ss_a.append(results)
			if(cell>6):
				results =c[cell][e]
				ss_c.append(results)
		ss_a.extend(ss_c)
		ws7.append(ss_a)	
	##########Group2-Group3##############	
	diff1 = list(set(b_1) & (set(c_1)) - (set(a_1)))
	le7=len(diff1)
	for row11 in diff1:
		ss_b=[]
		ss_c=[]
		d=b[0].index(row11)
		e=c[0].index(row11)
		for cell in range(0,(sheet.max_column)):
			results =b[cell][d]
			ss_b.append(results)
			if(cell>6):
				results =c[cell][e]
				ss_c.append(results)
		ss_b.extend(ss_c)
		ws8.append(ss_b)
	##########title##############
	ws1.merge_cells('A1:G2')
	ws2.merge_cells('A1:G2')
	ws3.merge_cells('A1:G1')
	ws4.merge_cells('A1:G1')
	ws5.merge_cells('A1:G1')
	ws6.merge_cells('A1:G2')
	ws7.merge_cells('A1:G2')
	ws8.merge_cells('A1:G2')
	
	ws1.cell(row=2,column=8).value='FPKM'
	ws2.cell(row=2,column=8).value='FPKM'
	ws3.cell(row=1,column=8).value='FPKM'
	ws4.cell(row=1,column=8).value='FPKM'
	ws5.cell(row=1,column=8).value='FPKM'
	ws6.cell(row=2,column=8).value='FPKM'
	ws7.cell(row=2,column=8).value='FPKM'
	ws8.cell(row=2,column=8).value='FPKM'
	
	ws1.cell(row=2,column=8).font=ft1
	ws1.cell(row=2,column=8).alignment=ali
	ws2.cell(row=2,column=8).font=ft1
	ws2.cell(row=2,column=8).alignment=ali
	ws3.cell(row=1,column=8).font=ft1
	ws3.cell(row=1,column=8).alignment=ali
	ws4.cell(row=1,column=8).font=ft1
	ws4.cell(row=1,column=8).alignment=ali
	ws5.cell(row=1,column=8).font=ft1
	ws5.cell(row=1,column=8).alignment=ali
	ws6.cell(row=2,column=8).font=ft1
	ws6.cell(row=2,column=8).alignment=ali
	ws7.cell(row=2,column=8).font=ft1
	ws7.cell(row=2,column=8).alignment=ali
	ws8.cell(row=2,column=8).font=ft1
	ws8.cell(row=2,column=8).alignment=ali
	
	ws1.merge_cells('H2:I2')
	ws2.merge_cells('H2:I2')
	ws3.merge_cells('H1:I1')
	ws4.merge_cells('H1:I1')
	ws5.merge_cells('H1:I1')
	ws6.merge_cells('H2:I2')
	ws7.merge_cells('H2:I2')
	ws8.merge_cells('H2:I2')
	
	ws1.cell(row=2,column=10).value='Fold Change'
	ws2.cell(row=2,column=10).value='Fold Change'
	ws3.cell(row=1,column=10).value='Fold Change'
	ws4.cell(row=1,column=10).value='Fold Change'
	ws5.cell(row=1,column=10).value='Fold Change'
	ws6.cell(row=2,column=10).value='Fold Change'
	ws7.cell(row=2,column=10).value='Fold Change'
	ws8.cell(row=2,column=10).value='Fold Change'
	
	ws1.cell(row=2,column=10).font=ft1
	ws1.cell(row=2,column=10).alignment=ali
	ws2.cell(row=2,column=10).font=ft1
	ws2.cell(row=2,column=10).alignment=ali
	ws3.cell(row=1,column=10).font=ft1
	ws3.cell(row=1,column=10).alignment=ali
	ws4.cell(row=1,column=10).font=ft1
	ws4.cell(row=1,column=10).alignment=ali
	ws5.cell(row=1,column=10).font=ft1
	ws5.cell(row=1,column=10).alignment=ali
	ws6.cell(row=2,column=10).font=ft1
	ws6.cell(row=2,column=10).alignment=ali
	ws7.cell(row=2,column=10).font=ft1
	ws7.cell(row=2,column=10).alignment=ali
	ws8.cell(row=2,column=10).font=ft1
	ws8.cell(row=2,column=10).alignment=ali
	
	ws1.merge_cells('J2:K2')
	ws2.merge_cells('J2:K2')
	ws3.merge_cells('J1:K1')
	ws4.merge_cells('J1:K1')
	ws5.merge_cells('J1:K1')
	ws6.merge_cells('J2:K2')
	ws7.merge_cells('J2:K2')
	ws8.merge_cells('J2:K2')
	
	ws1.cell(row=2,column=12).value='Statistic'
	ws2.cell(row=2,column=12).value='Statistic'
	ws3.cell(row=1,column=12).value='Statistic'
	ws4.cell(row=1,column=12).value='Statistic'
	ws5.cell(row=1,column=12).value='Statistic'
	ws6.cell(row=2,column=12).value='Statistic'
	ws7.cell(row=2,column=12).value='Statistic'
	ws8.cell(row=2,column=12).value='Statistic'
	
	ws1.cell(row=2,column=12).font=ft1
	ws1.cell(row=2,column=12).alignment=ali
	ws2.cell(row=2,column=12).font=ft1
	ws2.cell(row=2,column=12).alignment=ali
	ws3.cell(row=1,column=12).font=ft1
	ws3.cell(row=1,column=12).alignment=ali
	ws4.cell(row=1,column=12).font=ft1
	ws4.cell(row=1,column=12).alignment=ali
	ws5.cell(row=1,column=12).font=ft1
	ws5.cell(row=1,column=12).alignment=ali
	ws6.cell(row=2,column=12).font=ft1
	ws6.cell(row=2,column=12).alignment=ali
	ws7.cell(row=2,column=12).font=ft1
	ws7.cell(row=2,column=12).alignment=ali
	ws8.cell(row=2,column=12).font=ft1
	ws8.cell(row=2,column=12).alignment=ali
	
	ws1.merge_cells('L2:N2')
	ws2.merge_cells('L2:N2')
	ws3.merge_cells('L1:N1')
	ws4.merge_cells('L1:N1')
	ws5.merge_cells('L1:N1')
	ws6.merge_cells('L2:N2')
	ws7.merge_cells('L2:N2')
	ws8.merge_cells('L2:N2')
	
	ws1.cell(row=2,column=15).value='FPKM'
	ws2.cell(row=2,column=15).value='FPKM'
	ws6.cell(row=2,column=15).value='FPKM'
	ws7.cell(row=2,column=15).value='FPKM'
	ws8.cell(row=2,column=15).value='FPKM'
	
	ws1.cell(row=2,column=15).font=ft1
	ws1.cell(row=2,column=15).alignment=ali
	ws2.cell(row=2,column=15).font=ft1
	ws2.cell(row=2,column=15).alignment=ali
	ws6.cell(row=2,column=15).font=ft1
	ws6.cell(row=2,column=15).alignment=ali
	ws7.cell(row=2,column=15).font=ft1
	ws7.cell(row=2,column=15).alignment=ali
	ws8.cell(row=2,column=15).font=ft1
	ws8.cell(row=2,column=15).alignment=ali
	
	ws1.merge_cells('O2:P2')
	ws2.merge_cells('O2:P2')
	ws6.merge_cells('O2:P2')
	ws7.merge_cells('O2:P2')
	ws8.merge_cells('O2:P2')
	
	ws1.cell(row=2,column=17).value='Fold Change'
	ws2.cell(row=2,column=17).value='Fold Change'
	ws6.cell(row=2,column=17).value='Fold Change'
	ws7.cell(row=2,column=17).value='Fold Change'
	ws8.cell(row=2,column=17).value='Fold Change'
	
	ws1.cell(row=2,column=17).font=ft1
	ws1.cell(row=2,column=17).alignment=ali
	ws2.cell(row=2,column=17).font=ft1
	ws2.cell(row=2,column=17).alignment=ali
	ws6.cell(row=2,column=17).font=ft1
	ws6.cell(row=2,column=17).alignment=ali
	ws7.cell(row=2,column=17).font=ft1
	ws7.cell(row=2,column=17).alignment=ali
	ws8.cell(row=2,column=17).font=ft1
	ws8.cell(row=2,column=17).alignment=ali
	
	ws1.merge_cells('Q2:R2')
	ws2.merge_cells('Q2:R2')
	ws6.merge_cells('Q2:R2')
	ws7.merge_cells('Q2:R2')
	ws8.merge_cells('Q2:R2')
	
	ws1.cell(row=2,column=19).value='Statistic'
	ws2.cell(row=2,column=19).value='Statistic'
	ws6.cell(row=2,column=19).value='Statistic'
	ws7.cell(row=2,column=19).value='Statistic'
	ws8.cell(row=2,column=19).value='Statistic'
	
	ws1.cell(row=2,column=19).font=ft1
	ws1.cell(row=2,column=19).alignment=ali
	ws2.cell(row=2,column=19).font=ft1
	ws2.cell(row=2,column=19).alignment=ali
	ws6.cell(row=2,column=19).font=ft1
	ws6.cell(row=2,column=19).alignment=ali
	ws7.cell(row=2,column=19).font=ft1
	ws7.cell(row=2,column=19).alignment=ali
	ws8.cell(row=2,column=19).font=ft1
	ws8.cell(row=2,column=19).alignment=ali
	
	ws1.merge_cells('S2:U2')
	ws2.merge_cells('S2:U2')
	ws6.merge_cells('S2:U2')
	ws7.merge_cells('S2:U2')
	ws8.merge_cells('S2:U2')
	
	ws1.cell(row=1,column=8).value=n_1
	ws2.cell(row=1,column=8).value=n_1
	ws6.cell(row=1,column=8).value=n_1
	ws7.cell(row=1,column=8).value=n_1
	ws8.cell(row=1,column=8).value=n_2
	
	ws1.cell(row=1,column=8).font=ft1
	ws1.cell(row=1,column=8).alignment=ali
	ws2.cell(row=1,column=8).font=ft1
	ws2.cell(row=1,column=8).alignment=ali
	ws6.cell(row=1,column=8).font=ft1
	ws6.cell(row=1,column=8).alignment=ali
	ws7.cell(row=1,column=8).font=ft1
	ws7.cell(row=1,column=8).alignment=ali
	ws8.cell(row=1,column=8).font=ft1
	ws8.cell(row=1,column=8).alignment=ali
	
	ws1.merge_cells('H1:N1')
	ws2.merge_cells('H1:N1')
	ws6.merge_cells('H1:N1')
	ws7.merge_cells('H1:N1')
	ws8.merge_cells('H1:N1')
	
	
	ws1.cell(row=1,column=15).value=n_2
	ws2.cell(row=1,column=15).value=n_2
	ws6.cell(row=1,column=15).value=n_2
	ws7.cell(row=1,column=15).value=n_3
	ws8.cell(row=1,column=15).value=n_3
	
	ws1.cell(row=1,column=15).font=ft1
	ws1.cell(row=1,column=15).alignment=ali
	ws2.cell(row=1,column=15).font=ft1
	ws2.cell(row=1,column=15).alignment=ali
	ws6.cell(row=1,column=15).font=ft1
	ws6.cell(row=1,column=15).alignment=ali
	ws7.cell(row=1,column=15).font=ft1
	ws7.cell(row=1,column=15).alignment=ali
	ws8.cell(row=1,column=15).font=ft1
	ws8.cell(row=1,column=15).alignment=ali
	
	ws1.merge_cells('O1:U1')
	ws2.merge_cells('O1:U1')
	ws6.merge_cells('O1:U1')
	ws7.merge_cells('O1:U1')
	ws8.merge_cells('O1:U1')
	
	ws1.cell(row=1,column=22).value=n_3
	ws2.cell(row=1,column=22).value=n_3
	
	ws1.cell(row=1,column=22).font=ft1
	ws1.cell(row=1,column=22).alignment=ali
	ws2.cell(row=1,column=22).font=ft1
	ws2.cell(row=1,column=22).alignment=ali
	
	ws1.merge_cells('V1:AB1')
	ws2.merge_cells('V1:AB1')
	
	ws1.cell(row=2,column=22).value='FPKM'
	ws2.cell(row=2,column=22).value='FPKM'
	
	ws1.cell(row=2,column=22).font=ft1
	ws1.cell(row=2,column=22).alignment=ali
	ws2.cell(row=2,column=22).font=ft1
	ws2.cell(row=2,column=22).alignment=ali
	
	ws1.merge_cells('V2:W2')
	ws2.merge_cells('V2:W2')
	
	ws1.cell(row=2,column=24).value='Fold Change'
	ws2.cell(row=2,column=24).value='Fold Change'
	
	ws1.cell(row=2,column=24).font=ft1
	ws1.cell(row=2,column=24).alignment=ali
	ws2.cell(row=2,column=24).font=ft1
	ws2.cell(row=2,column=24).alignment=ali
	
	ws1.merge_cells('X2:Y2')
	ws2.merge_cells('X2:Y2')
	
	ws1.cell(row=2,column=26).value='Statistic'
	ws2.cell(row=2,column=26).value='Statistic'
	
	ws1.cell(row=2,column=26).font=ft1
	ws1.cell(row=2,column=26).alignment=ali
	ws2.cell(row=2,column=26).font=ft1
	ws2.cell(row=2,column=26).alignment=ali
	
	ws1.merge_cells('Z2:AB2')
	ws2.merge_cells('Z2:AB2')
	
	
	
	##########Venn##############
	venn3([set(a_1),set(b_1),set(c_1)], ('Group1', 'Group2', 'Group3'))
	c=os_s+'\\'+n_1+"_"+n_2+"_"+n_3+'Vennplot.png'
	plt.savefig(c, dpi=800)
	plt.cla()
	example = from_memberships([[],['Group1'],['Group2'],['Group3'],['Group1','Group2'],['Group1','Group3'],['Group2','Group3'],['Group1','Group2','Group3']],
								data=[0,le2,le3,le4,le5,le6,le7,le1])
	plot(example)
	cc=os_s+'\\'+n_1+"_"+n_2+"_"+n_3+'UpSetplot.png'
	plt.savefig(cc, dpi=800)
	
	
	ws4 = wo["Sheet"]
	ws4.title="Venn"
	img = Image(c)
	ws4.add_image(img)
	newsize = (800, 800)
	img.width, img.height = newsize
	
	img = Image(cc)
	ws4.add_image(img,'N1')
	newsize = (800, 800)
	img.width, img.height = newsize
	
	d=os_s+'\\'+n_1+"_"+n_2+"_"+n_3+".xlsx"
	wo.save(d)
	plt.show()


def callback():
	text = askopenfilename(filetype = (("xlsx files", "*.xlsx"),("","")))
	csvfile=text
	return path.set(csvfile)
	
def callback1():
	text = askopenfilename(filetype = (("xlsx files", "*.xlsx"),("","")))
	csvfile=text
	return path1.set(csvfile)	

def callback2(csvfile3="csvfile"):
	text1 = askopenfilename(filetype = (("xlsx files", "*.xlsx"),("","")))
	csvfile1=text1
	SaveDirectory = os.getcwd()
	
	b1=re.search('/(\w+)\.',csvfile1).group(1)
	b=re.search('/(\w+)\.',path.get()).group(1)
	b0=re.search('/(\w+)\.',path1.get()).group(1)
	
	venn_d(xlsx1=path.get(),xlsx2=path1.get(),xlsx3=csvfile1,os_s=SaveDirectory,n_1=b,n_2=b0,n_3=b1)

root = Tk()
root.title('Venn diagram')
path = StringVar()
Label(root,text = "目標路徑(Group_1):").grid(row = 0, column = 0)
Entry(root, textvariable = path).grid(row = 0, column = 1)
Button(root, text = "路徑選擇", command = callback).grid(row = 0, column = 2)

path1 = StringVar()
Label(root,text = "目標路徑(Group_2):").grid(row = 1, column = 0)
Entry(root, textvariable = path1).grid(row = 1, column = 1)
Button(root, text = "路徑選擇", command = callback1).grid(row = 1, column = 2)

path2 = StringVar()
Label(root,text = "目標路徑(Group_3):").grid(row = 2, column = 0)
Entry(root, textvariable = path2).grid(row = 2, column = 1)
Button(root, text = "路徑選擇", command = callback2).grid(row = 2, column = 2)

root.mainloop()
