import numpy as np
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
import networkx as nx
import seaborn as sns
import warnings; warnings.filterwarnings(action='once')
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import *
from tkinter.filedialog import askopenfilename
import re
import os
import random

def network(xlsx2="Example_FC2X.xlsx",os_s="c",lent2=2):
	wb = load_workbook(xlsx2)
	sheets = wb.sheetnames
	sheets1=wb.get_sheet_by_name("ID")
	l_s=len(sheets)
	data=sheets1.values
	colums=next(data)[0:]
	df = pd.DataFrame(data, columns=colums)
	dict_1 = {}
	for i in range(0,df.shape[0]):
		dict_1[df["Gene name"][i]]=df["log2FC"][i]

	for ii in range(0,(l_s-1)):
		sheets1=wb.get_sheet_by_name(sheets[ii])
		data=sheets1.values
		colums=next(data)[0:]
		df = pd.DataFrame(data, columns=colums)
		desc=[]
		list2 =[]
		lent=df.shape[0]
		d1={}
		ddd=[]
		node_sizes=[]
		if (df.shape[0]>lent2):
			lent=lent2
		for k in range(0,lent):
			tex=df["geneID"][k]
			desc.append(df["Description"][k])
			x1=tex.split('/')
			node_sizes.append(len(x1)*100)
			d1.setdefault(df["Description"][k],[]).append(x1)
			for l in list(x1):
				list2.append(l)
		list2_2 = list(set(list2))
		list1=np.hstack((desc,list2_2))

		G=nx.Graph()
		#G=nx.DiGraph()
		for i in list(list1):
			G.add_node(i)

		for i in list(desc):
			try:
				print(i)
				if( len(d1[i][0])>0):
					for j in list(d1[i][0]):
						G.add_edge(i,j)
			except Exception as e:
				pass
			continue

		pos = nx.layout.spring_layout(G)
		node_sizes1 = []
		#for i in range(len(G)-lent):
		#	node_sizes1.append(200)
		#node_sizes_ok=np.hstack((node_sizes,node_sizes1))
		M = G.number_of_edges()
		edge_colors = range(2, M + 2)
		#edge_alphas = [(5 + i) / (M + 4) for i in range(M)]
		cmap1 = sns.diverging_palette(133, 10, n=256, sep=64, as_cmap=True, center="light")
		color1=[]
		plt.figure(figsize=(28,15))
		for i in list(list2_2):
			try:
				color1.append(dict_1[i])
			except Exception as e:
				color1.append(0)
				pass
			continue
		nodes = nx.draw_networkx_nodes(G, pos,nodelist=list1[0:lent], node_size=node_sizes,node_color="lightyellow")
		nodes1 = nx.draw_networkx_nodes(G, pos,nodelist=list1[lent:], node_size=300,node_color=color1,cmap=cmap1, vmin=-3, vmax=3)
		edges = nx.draw_networkx_edges(G, pos, arrowstyle='->',arrowsize=10, edge_color="#DDDDDD", width=2)
		
		pos_attrs = {}
		for node, coords in pos.items():
			pos_attrs[node]=(coords[0],coords[1]+0.01)
		
		nx.draw_networkx_labels(G, pos_attrs, fontsize=6)

		plt.colorbar(nodes1)
		ax = plt.gca()
		ax.set_axis_off()
		
		plt.title("Network plot", fontsize=35)
		
		name1=sheets[ii]
		b2=re.search('(_\w+)',name1).group(1)
		c1=os_s+b2
		plt.gcf()#.set_size_inches(30, 15)
		
		plt.savefig(c1+'_network_plot.png', format='png', bbox_inches='tight', dpi=500)
		plt.close('all')
		print(sheets[ii])
	plt.show()

def callback2(k=2):
	text1 = askopenfilename(filetype = (("xlsx files", "*.xlsx"),("","")))
	csvfile1=text1
	SaveDirectory = os.getcwd()
	b1=re.search('/(\w+)\.',csvfile1).group(1)
	c=SaveDirectory+'\\'+b1
	network(xlsx2=text1,os_s=c,lent2=number_thr.get())
	
root = Tk()
root.title('Network')

number_thr = IntVar()
number_thr.set(2)
Label(root, text='number of pathway: ').grid(row = 0, column = 0)
Entry(root, textvariable = number_thr).grid(row = 0, column = 1)

path1 = StringVar()
Label(root,text = "目標路徑:").grid(row = 1, column = 0)
Entry(root, textvariable = path1).grid(row = 1, column = 1)
Button(root, text = "路徑選擇", command = callback2).grid(row = 1, column = 2)

root.mainloop()