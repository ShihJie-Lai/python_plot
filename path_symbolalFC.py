import numpy as np
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
import networkx as nx
import seaborn as sns
import warnings; warnings.filterwarnings(action='once')
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors, Font, Fill, NamedStyle,Color
from tkinter import *
from tkinter.filedialog import askopenfilename
import re
import os
import random

def network(xlsx2="Example_FC2X.xlsx",os_s="c",lent2=2):
	wb = load_workbook(xlsx2)
	sheets = wb.sheetnames

	sheets1=wb.get_sheet_by_name("ID")
	l_s=len(sheets)
	data=sheets1.values
	colums=next(data)[0:]
	df = pd.DataFrame(data, columns=colums)
	dict_1 = {}
	dict_2 = {}
	
	#####
	wo = Workbook()	
	ali = Alignment(horizontal='center', vertical='center')
	ft1=Font(name='Arial',color=colors.RED, size=12,b=True)
	####
	pattern = r".+(GSEA).+"
	
	for i in range(0,df.shape[0]):
		dict_1[df["Gene name"][i]]=df["log2FC"][i]
		dict_2[df["Gene name"][i]]=df["FC"][i]

	for ii in range(0,(l_s-1)):
		ws1 = wo.create_sheet(sheets[ii],index=0)
		sheets1=wb.get_sheet_by_name(sheets[ii])
		data=sheets1.values
		colums=next(data)[0:]
		df = pd.DataFrame(data, columns=colums)
		desc=[]
		list2 =[]
		lent=df.shape[0]
		d1={}
		row_s=2
		if (df.shape[0]>lent2):
			lent=lent2
		if re.fullmatch(pattern, xlsx2):
			for k in range(0,lent):
				tex=df["core_enrichment"][k]
				x1=tex.split('/')
				st=()
				for kk in  range(0,len(x1)):
					if (kk==0):
						st=row_s
						ws1.cell(row=row_s,column=0+1).value=df["ID"][k]
						ws1.cell(row=row_s,column=0+1).alignment=ali	
						ws1.cell(row=row_s,column=1+1).value=df["Description"][k]
						ws1.cell(row=row_s,column=1+1).alignment=ali	
						ws1.cell(row=row_s,column=2+1).value=df["setSize"][k]
						ws1.cell(row=row_s,column=2+1).alignment=ali	
						ws1.cell(row=row_s,column=3+1).value=df["enrichmentScore"][k]
						ws1.cell(row=row_s,column=3+1).alignment=ali	
						ws1.cell(row=row_s,column=4+1).value=df["NES"][k]
						ws1.cell(row=row_s,column=4+1).alignment=ali	
						ws1.cell(row=row_s,column=5+1).value=df["pvalue"][k]
						ws1.cell(row=row_s,column=5+1).alignment=ali	
						ws1.cell(row=row_s,column=6+1).value=df["p.adjust"][k]
						ws1.cell(row=row_s,column=6+1).alignment=ali	
						ws1.cell(row=row_s,column=7+1).value=df["qvalues"][k]
						ws1.cell(row=row_s,column=7+1).alignment=ali
						ws1.cell(row=row_s,column=8+1).value=df["rank"][k]
						ws1.cell(row=row_s,column=8+1).alignment=ali
						ws1.cell(row=row_s,column=9+1).value=df["leading_edge"][k]
						ws1.cell(row=row_s,column=9+1).alignment=ali
						ws1.cell(row=row_s,column=10+1).value=x1[kk]
						if(x1[kk] in dict_2):
							ws1.cell(row=row_s,column=11+1).value=dict_2[x1[kk]]
							ws1.cell(row=row_s,column=12+1).value=dict_1[x1[kk]]
						else:
							ws1.cell(row=row_s,column=11+1).value="-"
							ws1.cell(row=row_s,column=12+1).value="-"
					else:
						ws1.cell(row=row_s,column=0+1).value=""
						ws1.cell(row=row_s,column=1+1).value=""
						ws1.cell(row=row_s,column=2+1).value=""
						ws1.cell(row=row_s,column=3+1).value=""
						ws1.cell(row=row_s,column=4+1).value=""
						ws1.cell(row=row_s,column=5+1).value=""
						ws1.cell(row=row_s,column=6+1).value=""
						ws1.cell(row=row_s,column=7+1).value=""
						ws1.cell(row=row_s,column=8+1).value=""
						ws1.cell(row=row_s,column=9+1).value=""
						ws1.cell(row=row_s,column=10+1).value=x1[kk]
						if( x1[kk] in dict_2):
							ws1.cell(row=row_s,column=11+1).value=dict_2[x1[kk]]
							ws1.cell(row=row_s,column=12+1).value=dict_1[x1[kk]]
						else:
							ws1.cell(row=row_s,column=11+1).value="-"
							ws1.cell(row=row_s,column=12+1).value="-"
					row_s=row_s+1
				ws1.merge_cells(start_row=st, start_column=1, end_row=row_s-1, end_column=1)
				ws1.merge_cells(start_row=st, start_column=2, end_row=row_s-1, end_column=2)
				ws1.merge_cells(start_row=st, start_column=3, end_row=row_s-1, end_column=3)
				ws1.merge_cells(start_row=st, start_column=4, end_row=row_s-1, end_column=4)
				ws1.merge_cells(start_row=st, start_column=5, end_row=row_s-1, end_column=5)
				ws1.merge_cells(start_row=st, start_column=6, end_row=row_s-1, end_column=6)
				ws1.merge_cells(start_row=st, start_column=7, end_row=row_s-1, end_column=7)
				ws1.merge_cells(start_row=st, start_column=8, end_row=row_s-1, end_column=8)
				ws1.merge_cells(start_row=st, start_column=9, end_row=row_s-1, end_column=9)
				ws1.merge_cells(start_row=st, start_column=10, end_row=row_s-1, end_column=10)
				ws1.cell(row=1,column=1).value="ID"
				ws1.cell(row=1,column=1).font=ft1
				ws1.cell(row=1,column=2).value="Description"
				ws1.cell(row=1,column=2).font=ft1
				ws1.cell(row=1,column=3).value="setSize"
				ws1.cell(row=1,column=3).font=ft1
				ws1.cell(row=1,column=4).value="enrichmentScore"
				ws1.cell(row=1,column=4).font=ft1
				ws1.cell(row=1,column=5).value="NES"
				ws1.cell(row=1,column=5).font=ft1
				ws1.cell(row=1,column=6).value="pvalue"
				ws1.cell(row=1,column=6).font=ft1
				ws1.cell(row=1,column=7).value="p.adjust"
				ws1.cell(row=1,column=7).font=ft1
				ws1.cell(row=1,column=8).value="qvalues"
				ws1.cell(row=1,column=8).font=ft1
				ws1.cell(row=1,column=9).value="rank"
				ws1.cell(row=1,column=9).font=ft1
				ws1.cell(row=1,column=10).value="leading_edge"
				ws1.cell(row=1,column=10).font=ft1
				ws1.cell(row=1,column=11).value="core_enrichment"
				ws1.cell(row=1,column=11).font=ft1
				ws1.cell(row=1,column=12).value="FC"
				ws1.cell(row=1,column=12).font=ft1
				ws1.cell(row=1,column=13).value="log2FC"
				ws1.cell(row=1,column=13).font=ft1
		else:
			for k in range(0,lent):
				tex=df["geneID"][k]
				x1=tex.split('/')
				st=()
				for kk in  range(0,len(x1)):
					if (kk==0):
						st=row_s
						ws1.cell(row=row_s,column=0+1).value=df["ID"][k]
						ws1.cell(row=row_s,column=0+1).alignment=ali	
						ws1.cell(row=row_s,column=1+1).value=df["Description"][k]
						ws1.cell(row=row_s,column=1+1).alignment=ali	
						ws1.cell(row=row_s,column=2+1).value=df["GeneRatio"][k]
						ws1.cell(row=row_s,column=2+1).alignment=ali	
						ws1.cell(row=row_s,column=3+1).value=df["BgRatio"][k]
						ws1.cell(row=row_s,column=3+1).alignment=ali	
						ws1.cell(row=row_s,column=4+1).value=df["pvalue"][k]
						ws1.cell(row=row_s,column=4+1).alignment=ali	
						ws1.cell(row=row_s,column=5+1).value=df["p.adjust"][k]
						ws1.cell(row=row_s,column=5+1).alignment=ali	
						ws1.cell(row=row_s,column=6+1).value=df["qvalue"][k]
						ws1.cell(row=row_s,column=6+1).alignment=ali	
						ws1.cell(row=row_s,column=7+1).value=df["Count"][k]
						ws1.cell(row=row_s,column=7+1).alignment=ali	
						ws1.cell(row=row_s,column=8+1).value=x1[kk]
						if(x1[kk] in dict_2):
							ws1.cell(row=row_s,column=9+1).value=dict_2[x1[kk]]
							ws1.cell(row=row_s,column=10+1).value=dict_1[x1[kk]]
						else:
							ws1.cell(row=row_s,column=9+1).value="-"
							ws1.cell(row=row_s,column=10+1).value="-"
					else:
						ws1.cell(row=row_s,column=0+1).value=""
						ws1.cell(row=row_s,column=1+1).value=""
						ws1.cell(row=row_s,column=2+1).value=""
						ws1.cell(row=row_s,column=3+1).value=""
						ws1.cell(row=row_s,column=4+1).value=""
						ws1.cell(row=row_s,column=5+1).value=""
						ws1.cell(row=row_s,column=6+1).value=""
						ws1.cell(row=row_s,column=7+1).value=""
						ws1.cell(row=row_s,column=8+1).value=x1[kk]
						if(x1[kk] in dict_2):
							ws1.cell(row=row_s,column=9+1).value=dict_2[x1[kk]]
							ws1.cell(row=row_s,column=10+1).value=dict_1[x1[kk]]
						else:
							ws1.cell(row=row_s,column=9+1).value="-"
							ws1.cell(row=row_s,column=10+1).value="-"
					row_s=row_s+1
				ws1.merge_cells(start_row=st, start_column=1, end_row=row_s-1, end_column=1)
				ws1.merge_cells(start_row=st, start_column=2, end_row=row_s-1, end_column=2)
				ws1.merge_cells(start_row=st, start_column=3, end_row=row_s-1, end_column=3)
				ws1.merge_cells(start_row=st, start_column=4, end_row=row_s-1, end_column=4)
				ws1.merge_cells(start_row=st, start_column=5, end_row=row_s-1, end_column=5)
				ws1.merge_cells(start_row=st, start_column=6, end_row=row_s-1, end_column=6)
				ws1.merge_cells(start_row=st, start_column=7, end_row=row_s-1, end_column=7)
				ws1.merge_cells(start_row=st, start_column=8, end_row=row_s-1, end_column=8)
				ws1.cell(row=1,column=1).value="ID"
				ws1.cell(row=1,column=1).font=ft1
				ws1.cell(row=1,column=2).value="Description"
				ws1.cell(row=1,column=2).font=ft1
				ws1.cell(row=1,column=3).value="GeneRatio"
				ws1.cell(row=1,column=3).font=ft1
				ws1.cell(row=1,column=4).value="BgRatio"
				ws1.cell(row=1,column=4).font=ft1
				ws1.cell(row=1,column=5).value="pvalue"
				ws1.cell(row=1,column=5).font=ft1
				ws1.cell(row=1,column=6).value="p.adjust"
				ws1.cell(row=1,column=6).font=ft1
				ws1.cell(row=1,column=7).value="qvalue"
				ws1.cell(row=1,column=7).font=ft1
				ws1.cell(row=1,column=8).value="Count"
				ws1.cell(row=1,column=8).font=ft1
				ws1.cell(row=1,column=9).value="geneID"
				ws1.cell(row=1,column=9).font=ft1
				ws1.cell(row=1,column=10).value="FC"
				ws1.cell(row=1,column=10).font=ft1
				ws1.cell(row=1,column=11).value="log2FC"
				ws1.cell(row=1,column=11).font=ft1
	d=os_s+"_FC.xlsx"
	print(d)
	print("OK!!!!!!")	
	wo.save(d)

def callback2(k=2):
	text1 = askopenfilename(filetype = (("xlsx files", "*.xlsx"),("","")))
	csvfile1=text1
	SaveDirectory = os.getcwd()
	b1=re.search('/(\w+)\.',csvfile1).group(1)
	c=SaveDirectory+'\\'+b1
	network(xlsx2=text1,os_s=c,lent2=number_thr.get())
	
root = Tk()
root.title('Network')

number_thr = IntVar()
number_thr.set(2)
Label(root, text='number of pathway: ').grid(row = 0, column = 0)
Entry(root, textvariable = number_thr).grid(row = 0, column = 1)

path1 = StringVar()
Label(root,text = "目標路徑:").grid(row = 1, column = 0)
Entry(root, textvariable = path1).grid(row = 1, column = 1)
Button(root, text = "路徑選擇", command = callback2).grid(row = 1, column = 2)

root.mainloop()