import numpy as np
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
import seaborn as sns
import warnings; warnings.filterwarnings(action='once')
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import *
from tkinter.filedialog import askopenfilename
import re
import os


def Pyramid(xlsx2="csvfile1",os_s="c"):
	wb = load_workbook(xlsx2)
	sheets = wb.sheetnames
	sheets1=wb.get_sheet_by_name("ID")
	data=sheets1.values
	colums=next(data)[0:]

	df = pd.DataFrame(data, columns=colums)
	dict_1 = {}

	for i in range(0,df.shape[0]):
		#print(df["Gene name"][i])
		dict_1[df["Gene name"][i]]=df["log2FC"][i]

	for j in range(0,4):
		desc=[]
		expression=[]
		count_c=[]
		sheets1=wb.get_sheet_by_name(sheets[j])
		data=sheets1.values
		colums=next(data)[0:]
		df = pd.DataFrame(data, columns=colums)
		lent=df.shape[0]
		if (df.shape[0]>10):
			lent=10
		for k in range(0,lent):
			tex=df["geneID"][k]
			x1=tex.split('/')
			up=0
			down=0
			for l in range(0,(len(x1))):
				try:
					test=x1[l]
					if(dict_1[x1[l]]>0):
						up=up+1
					else:
						down=down+1
				except Exception as e:
					pass
				continue
			desc.append(df["Description"][k])
			expression.append("up")
			count_c.append(up)
			desc.append(df["Description"][k])
			expression.append("down")
			count_c.append(-down)
		dict = {"Description": desc,"Expression": expression,"Count":count_c}
		select_df = pd.DataFrame(dict)
		print(select_df)
		plt.figure(figsize=(20,10), dpi= 200)
		group_col = 'Expression'
		order_of_bars = select_df.Description.unique()[::-1]
		colors = [plt.cm.Spectral(i/float(len(select_df[group_col].unique())-1)) for i in range(len(select_df[group_col].unique()))]
		print(colors)
		for c, group in zip(colors, select_df[group_col].unique()):
			sns.barplot(x='Count', y='Description', data=select_df.loc[select_df[group_col]==group, :], order=order_of_bars, color=c, label=group)
		plt.xlabel("Count")
		plt.ylabel("Description")
		plt.yticks(fontsize=12)
		plt.title("Pyramid plot", fontsize=22)
		plt.legend()
		############
		name1=sheets[j]
		b2=re.search('(_\w+)',name1).group(1)
		c1=os_s+b2
		print(c1)
		plt.savefig(c1+'_Pyramid_plot.png', format='png', bbox_inches='tight',  dpi=80)
		############
		#plt.savefig('pcaplot_3d.png', format='png', bbox_inches='tight',  dpi=80)
	plt.show()

		

def callback2(csvfile3="csvfile"):
	text1 = askopenfilename(filetype = (("xlsx files", "*.xlsx"),("","")))
	csvfile1=text1
	SaveDirectory = os.getcwd()
	b1=re.search('/(\w+)\.',csvfile1).group(1)
	c=SaveDirectory+'\\'+b1
	Pyramid(xlsx2=csvfile1,os_s=c)	

root = Tk()
root.title('Pyramid diagram')
path1 = StringVar()
Label(root,text = "目標路徑:").grid(row = 0, column = 0)
Entry(root, textvariable = path1).grid(row = 0, column = 1)
Button(root, text = "路徑選擇", command = callback2).grid(row = 0, column = 2)

root.mainloop()