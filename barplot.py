import numpy as np
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
import seaborn as sns
import warnings; warnings.filterwarnings(action='once')
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import *
from tkinter.filedialog import askopenfilename
import re
import os
import plotly_express as px
import plotly.graph_objects as go

def Pyramid(xlsx2="csvfile1",shape=10,col_thr=0,colup="gray",os_s="c"):
	wb = load_workbook(xlsx2)
	sheets = wb.sheetnames

	for j in range(0,2):
		desc=[]
		expression=[]
		count_c=[]
		pv=[]
		tit=""
		sheets1=wb.get_sheet_by_name(sheets[j])
		data=sheets1.values
		colums=next(data)[0:]
		df = pd.DataFrame(data, columns=colums)
		lent=df.shape[0]
		if (df.shape[0]>shape):
			lent=shape
		for k in range(0,lent):
			desc.append(df["Description"][k])
			expression.append(df[colums[0]][k])
			count_c.append(df["Count"][k])
			if (col_thr==0):
				pv.append(df["pvalue"][k])
				tit="pvalue"
			elif (col_thr==1):
				pv.append(df["p.adjust"][k])
				tit="p.adjust"
			elif (col_thr==2):
				pv.append(df["qvalue"][k])
				tit="qvalue"
			else :
				pv.append(0)
		print(lent)		
		dict1 = {"Description": desc,"Expression": expression,"Count":count_c,"pv":pv}
		select_df = pd.DataFrame(dict1)
		print(select_df)
		if(j==0):
			x=[expression,desc]
			fig = go.Figure()#layout=go.Layout(height=1500, width=1000
			fig.add_bar(x=x,y=count_c, marker=dict(color=pv,colorbar=dict(title=tit),colorscale=colup))
			fig.update_layout(annotations = [dict(
									x = 0,
									y = 1.1,
									font = dict(size = 50))],
									font_size=20,
									legend_font_size=40,
									xaxis_tickfont_size=20,
									yaxis=dict(title='gene counts',titlefont_size=22,tickfont_size=30,showgrid=False),
									height=1500,width=2000)
			
		else:
			fig = px.bar(select_df, x='Description', y='Count', color='pv',height=1000,width=800,labels={"pv": tit},color_continuous_scale=colup)
			fig.update_layout(annotations = [dict(
									x = 0,
									y = 1.1,
									font = dict(size = 40))],
									font_size=20,
									xaxis_tickfont_size=20,
									yaxis=dict(title='gene counts',titlefont_size=22,tickfont_size=30,showgrid=False),
									height=1500,width=2000)
		############
		name1=sheets[j]
		print(name1)
		b2=re.search('(_\w+)',name1).group(1)
		c1=os_s+b2
		#fig.write_image(c1+'_bar_plot.svg')
		fig.write_html(c1+'_bar_plot.html')
		print(c1)


		#fig.savefig(c1+'_Pyramid_plot.png', format='png', bbox_inches='tight',  dpi=80)
		############
		#plt.savefig('pcaplot_3d.png', format='png', bbox_inches='tight',  dpi=80)


		

def callback2(csvfile3="csvfile"):
	text1 = askopenfilename(filetype = (("xlsx files", "*.xlsx"),("","")))
	csvfile1=text1
	SaveDirectory = os.getcwd()
	b1=re.search('/(\w+)\.',csvfile1).group(1)
	c=SaveDirectory+'\\'+b1
	Pyramid(xlsx2=csvfile1,shape=xmin.get(),col_thr=col_thr.get(),colup=colup.get(),os_s=c)	

root = Tk()
root.title('bar plot')


Label(root,text = "Color:").grid(row = 0, column = 0)

colup = StringVar() 
colup.set("gray") 
Radiobutton(root,text = 'black to white',variable = colup,value = "gray").grid(row = 0, column = 1)
Radiobutton(root,text = 'red to green',variable = colup,value = "rdylgn").grid(row = 0, column = 2)

Label(root,text = "Color scale:").grid(row = 1, column = 0)

col_thr = DoubleVar()
col_thr.set(0)
Radiobutton(root,text = 'p-value',variable = col_thr,value = 0).grid(row = 2, column = 0)
Radiobutton(root,text = 'p.adjust',variable = col_thr,value = 1).grid(row = 2, column = 1)
Radiobutton(root,text = 'q-value',variable = col_thr,value = 2).grid(row = 2, column = 2)
Radiobutton(root,text = 'None',variable = col_thr,value = 3).grid(row = 2, column = 3)


xmin = IntVar()
xmin.set(10)
Label(root, text='num: ').grid(row = 3, column = 0)
Entry(root, textvariable = xmin).grid(row = 3, column = 1)


path1 = StringVar()
Label(root,text = "目標路徑:").grid(row = 4, column = 0)
Entry(root, textvariable = path1).grid(row = 4, column = 1)
Button(root, text = "路徑選擇", command = callback2).grid(row = 4, column = 2)

root.mainloop()