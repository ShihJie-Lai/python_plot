import numpy as np
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
import networkx as nx
import seaborn as sns
import warnings; warnings.filterwarnings(action='once')
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import *
from tkinter.filedialog import askopenfilename
import re
import os
from upsetplot import plot
from upsetplot import from_contents


def network(xlsx2="Example_FC2X.xlsx",os_s="c",lent2=3):
	wb = load_workbook(xlsx2)
	sheets = wb.sheetnames

	for ii in range(0,4):
		sheets1=wb.get_sheet_by_name(sheets[ii])
		data=sheets1.values
		colums=next(data)[0:]
		df = pd.DataFrame(data, columns=colums)
		desc=[]
		list2 =[]
		lent=df.shape[0]
		d1={}
		ddd=[]
		node_sizes=[]
		if (df.shape[0]>lent2):
			lent=lent2
		for k in range(0,lent):
			tex=df["geneID"][k]
			desc.append(df["ID"][k])
			x1=tex.split('/')
			#node_sizes.append(len(x1)*100)
			d1.setdefault(df["ID"][k],[]).append(x1)
		#print(d1[df["ID"][0]][0])
		example=from_contents(d1)
		#print(example)
		plot(example)
		plt.title("Upset plot", fontsize=22)
		name1=sheets[ii]
		b2=re.search('(_\w+)',name1).group(1)
		c1=os_s+b2
		plt.gcf().set_size_inches(30, 20)
		plt.savefig(c1+'_upset_plot.png', format='png', bbox_inches='tight', dpi=200)
		plt.close('all')
	print("End")


def callback2(k=2):
	text1 = askopenfilename(filetype = (("xlsx files", "*.xlsx"),("","")))
	csvfile1=text1
	SaveDirectory = os.getcwd()
	b1=re.search('/(\w+)\.',csvfile1).group(1)
	c=SaveDirectory+'\\'+b1
	network(xlsx2=text1,os_s=c,lent2=number_thr.get())
	
root = Tk()
root.title('Network')

number_thr = IntVar()
number_thr.set(3)
Label(root, text='number of pathway: ').grid(row = 0, column = 0)
Entry(root, textvariable = number_thr).grid(row = 0, column = 1)

path1 = StringVar()
Label(root,text = "目標路徑:").grid(row = 1, column = 0)
Entry(root, textvariable = path1).grid(row = 1, column = 1)
Button(root, text = "路徑選擇", command = callback2).grid(row = 1, column = 2)

root.mainloop()